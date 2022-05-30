from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
import time
import os
from tqdm import tqdm

username = "na_anhson"
password = "Hieu@TTYT123"

ds_xa_huyenanhson_string =" thị trấn Anh Sơn , Bình Sơn, Cẩm Sơn, Cao Sơn, Đỉnh Sơn, Đức Sơn, Hoa Sơn, Hội Sơn, Hùng Sơn, Khai Sơn, Lạng Sơn, Lĩnh Sơn, Long Sơn, Phúc Sơn, Tam Sơn, Tào Sơn, Thạch Sơn, Thành Sơn, Thọ Sơn, Tường Sơn, Vĩnh Sơn"
ds_xa_huyenanhson = ds_xa_huyenanhson_string.split(",")
clear_ds_xa = [x.strip() for x in ds_xa_huyenanhson]


driver = webdriver.Firefox()
driver.get("https://tiemchungcovid19.moh.gov.vn/")
time.sleep(5)

username_input = driver.find_element_by_id("username")
password_input = driver.find_element_by_id("password")
submit_btn = driver.find_element_by_id("btnLogin")

username_input.send_keys(username)
password_input.send_keys(password)
submit_btn.click()

time.sleep(5)
otp = input("Nhập OTP: ")

otp_input = driver.find_element_by_id("otpTxt")
otp_input.send_keys(otp)
submit_btn = driver.find_element_by_xpath("/html/body/div/div/div/div/button[1]")
submit_btn.click()
time.sleep(5)

xa = 'thitran'
dest_filename = 'Thong_ke_xac_minh_doi_tuong_' + xa + ' - Copy.xlsx'
tmp_filename = 'copy_tmp'+ xa + ' - Copy.xlsx'
backup_filename = 'copy_backup'+ xa + ' - Copy.xlsx'
wb = load_workbook(filename = dest_filename)
ws = wb.active
for i in tqdm(range(7, ws.max_row+1)):
    try:
        cccd_key = ws.cell(i, 4)
        if (cccd_key.value != None) and (cccd_key.value != "Không có thông tin"):
            driver.get("https://tiemchungcovid19.moh.gov.vn/TiemChung/DoiTuong/IndexCovid")
            time.sleep(1)

            driver.find_element_by_xpath('/html/body/section/div[2]/div[3]/div[1]/div[1]/div/div[1]/table/tbody/tr/td[2]/button').click()
            search_name = ws.cell(i,5).value #Ttên
            search_phone = ws.cell(i, 10).value # SĐT
            search_name_input = driver.find_element_by_id('txtHoTenSearch')
            search_phone_input = driver.find_element_by_id('txtSoDienThoaiSearch')

            search_name_input.send_keys(search_name)
            search_phone_input.send_keys(search_phone)

            search_btn = driver.find_element_by_id('btnAdvancedSearch')
            clear = driver.find_element_by_xpath('//*[@id="select2-slDonViTao-container"]/span')
            clear.click()
            search_btn.click()
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="AdvancedSearchPopover"]/h3/span').click()

            link_edit = driver.find_element_by_xpath('//*[@id="doiTuongSearchResult"]/tbody/tr/td[2]/div')
            link_edit.click()
            time.sleep(1)

            btn_edit = driver.find_element_by_id('btnEdit')
            btn_edit.click()
            time.sleep(1)

            driver.find_element_by_xpath('//*[@id="thongTinCoBan"]/div[2]/div[2]/span/span[1]/span/span[2]').click()
            driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys("Việt Nam")
            driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys(Keys.ENTER)

            cccd_input = driver.find_element_by_id('txtCMT_Sua')
            cccd_input.clear()

            cccd_input.send_keys(cccd_key.value)

            current_xa = ws.cell(i,2).value

            if current_xa != "Không tìm thấy" and current_xa != None and current_xa != "nhiều hơn 1 kết quả":
                for xa in clear_ds_xa:
                    if xa in current_xa:
                        driver.find_element_by_xpath('//*[@id="thongTinCoBan"]/div[5]/div[1]/span/span[1]/span/span[2]').click()
                        driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys("Nghệ An")
                        driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys(Keys.ENTER)
                        driver.find_element_by_xpath('//*[@id="thongTinCoBan"]/div[5]/div[2]/span/span[1]/span/span[2]').click()
                        driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys("Anh Sơn")
                        driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys(Keys.ENTER)
                        driver.find_element_by_xpath('//*[@id="thongTinCoBan"]/div[5]/div[3]/span/span[1]/span/span[2]').click()
                        driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys(xa)
                        driver.find_element_by_xpath('/html/body/span/span/span[1]/input').send_keys(Keys.ENTER)
                        driver.find_element_by_id('btnApDungChoTamTru').click()

            save_btn = driver.find_element_by_id('btnSave')
            save_btn.click()
            time.sleep(1)
            try:
                duplicate = driver.find_element_by_xpath('/html/body/div[11]/div/div/div[1]/h4')
                ws.cell(i,1).value = "Duplicate"
            except:
                ws.cell(i,1).value = "Done"
        else:
            ws.cell(i,1).value = "Failed"
        wb.save(tmp_filename)
        os.rename(dest_filename, backup_filename)
        os.rename(tmp_filename, dest_filename)
        os.remove(backup_filename)
    except:
        print("Lỗi")

driver.quit()
