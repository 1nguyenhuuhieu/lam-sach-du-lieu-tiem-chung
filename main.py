from unittest import result
from selenium import webdriver
from openpyxl import load_workbook
import time
import os
from tqdm import tqdm

username = "na_as_ttytanhson"
password = "Jw3@b@qnPjj7vUT"


driver = webdriver.Chrome()
driver.get("http://hssk.kcb.vn/")
time.sleep(5)

username_input = driver.find_element_by_id("username")
password_input = driver.find_element_by_id("password")
submit_btn = driver.find_element_by_xpath("/html/body/app-root/app-login/div/div[2]/div/form/button")

username_input.send_keys(username)
password_input.send_keys(password)
submit_btn.click()

time.sleep(5)
otp = input("Nhập OTP: ")

otp_input = driver.find_element_by_id("otp")
otp_input.clear()
otp_input.send_keys(otp)
submit_btn = driver.find_element_by_xpath("/html/body/app-root/app-login/div/div[2]/div/div[3]/di/button")
submit_btn.click()
time.sleep(5)

xa = 'phucson'
dest_filename = 'Thong_ke_xac_minh_doi_tuong_' + xa + '.xlsx'
tmp_filename = 'tmp_' + xa + '.xlsx'
backup_filename = 'backup_' + xa + '.xlsx'
wb = load_workbook(filename = dest_filename)
ws = wb.active
for i in tqdm(range(7, ws.max_row+1)):
    driver.get("http://hssk.kcb.vn/#/ho-so-suc-khoe")
    time.sleep(1)
    try:
        advanced_search = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/div[3]/app-base-button[2]/button/span')
        advanced_search.click()
        name_input = driver.find_element_by_xpath('//*[@id="demo"]/div/div[5]/div/input')
        birth_input = driver.find_element_by_xpath('//*[@id="demo"]/div/div[17]/div/app-datepicker/div/input')
        search_btn = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/div[3]/app-base-button[1]/button')
        name = ws.cell(i,5).value #Họ và tên
        birth = ws.cell(i,6).value #Ngày sinh

        name_input.clear()
        name_input.send_keys(name)
        birth_input.clear()
        birth_input.send_keys(birth)
        search_btn.click()
        
        time.sleep(1)

        result_search = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/span[2]')
        # Nếu tìm thấy kết quả
        if result_search.text == "(1)":
            diachithuongtru_output = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/search-table/div/div/div[1]/table/tbody/tr/td[9]/div')
            phone_output = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/search-table/div/div/div[1]/table/tbody/tr/td[7]/div')
            madinhdanhyte_output = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/search-table/div/div/div[1]/table/tbody/tr/td[3]/div/a')
            ws.cell(i,2).value = diachithuongtru_output.text
            ws.cell(i,3).value = phone_output.text
            chitiet_btn = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-info-search/search-table/div/div/div[1]/table/tbody/tr[1]/td[11]/div/div[1]/span/button')
            chitiet_btn.click()
            time.sleep(1)
            cccd_output = driver.find_element_by_xpath('//*[@id="app-content"]/div[1]/app-nhan-khau-chi-tiet/div/div[2]/div/div[10]/span[2]')
            ws.cell(i,4).value = cccd_output.text
        elif result_search.text == "(0)":
            ws.cell(i,2).value = "Không tìm thấy"
        else:
            ws.cell(i,2).value = "nhiều hơn 1 kết quả"
        
        wb.save(tmp_filename)
        os.rename(dest_filename, backup_filename)
        os.rename(tmp_filename, dest_filename)
        os.remove(backup_filename)
    except:
        print("Lỗi")


driver.quit()
