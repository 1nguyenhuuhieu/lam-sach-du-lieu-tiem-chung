from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
import time
from tqdm import tqdm

username = "na_anhson"
password = "Hieu@TTYT123"


driver = webdriver.Chrome()
driver.get("https://tiemchungcovid19.moh.gov.vn/")
time.sleep(2)

username_input = driver.find_element_by_id("username")
password_input = driver.find_element_by_id("password")
submit_btn = driver.find_element_by_id("btnLogin")

username_input.send_keys(username)
password_input.send_keys(password)
submit_btn.click()

time.sleep(2)
otp = input("Nhập OTP: ")

otp_input = driver.find_element_by_id("otpTxt")
otp_input.send_keys(otp)
submit_btn = driver.find_element_by_xpath("/html/body/div/div/div/div/button[1]")
submit_btn.click()
time.sleep(2)

dest_filename = 'xoa.xlsx'
wb = load_workbook(filename = dest_filename)
ws = wb.active
for i in tqdm(range(6, ws.max_row+1)):
    try:
        driver.get("https://tiemchungcovid19.moh.gov.vn/TiemChung/DoiTuong/IndexCovid")
        time.sleep(1)

        driver.find_element_by_xpath('/html/body/section/div[2]/div[3]/div[1]/div[1]/div/div[1]/table/tbody/tr/td[2]/button').click()
        search_name = ws.cell(i,2).value #Ttên
        search_phone = ws.cell(i,7).value # SĐT
        search_name_input = driver.find_element_by_id('txtHoTenSearch')
        search_phone_input = driver.find_element_by_id('txtSoDienThoaiSearch')

        search_name_input.send_keys(search_name)
        search_phone_input.send_keys(search_phone)

        search_btn = driver.find_element_by_id('btnAdvancedSearch')
        clear = driver.find_element_by_xpath('//*[@id="select2-slDonViTao-container"]/span')
        clear.click()
        search_btn.click()
        time.sleep(1)
        link_edit = driver.find_element_by_xpath('//*[@id="doiTuongSearchResult"]/tbody/tr/td[2]/div')
        link_edit.click()
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="AdvancedSearchPopover"]/h3/span').click()
        driver.find_element_by_xpath('//*[@id="tab-user-info"]/li[2]/a').click()

        d = driver.execute_script("return document.getElementById('tblVacxin').rows.length;")

        if d==2:
            driver.find_element_by_xpath('//*[@id="tblVacxin"]/tbody/tr/td[8]/a[6]/i').click()
            driver.find_element_by_xpath('/html/body/div[9]/div/div/div[3]/button[2]').click()
        

        # try:
        #     driver.find_element_by_xpath('//*[@id="tblVacxin"]/tbody/tr[2]/td[3]')
        #     try:
        #         driver.find_element_by_xpath('//*[@id="tblVacxin"]/tbody/tr[1]/td[3]')
        #     except:
        #         driver.find_element_by_xpath('//*[@id="tblVacxin"]/tbody/tr[2]/td[8]/a[6]/i').click()
        #         print("Đã xóa")
        #         time.sleep(1)
        # except:
        #     pass
        time.sleep(1)

    except:
        print("Lỗi")

driver.quit()
